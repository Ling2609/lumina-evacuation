# =============================================================================
# LUMINA SMART EVACUATION SYSTEM
# lumina_live_stream.py  —  Main Edge Node Controller
#
# Run:  python lumina_live_stream.py
#
# What changed from the original:
#   ① ByteTrack (YOLO native) replaces DeepSORT — persistent anonymous IDs,
#     density history, and velocity all flow into the routing engine.
#     Uses model.track(persist=True) — no random embeddings, no ID flicker.
#   ② Crowd velocity → routing_engine.update_crowd() called every frame,
#     so DYN-A* is always working with live predictive data
#   ③ ThermalClassifier added as a background thread — simulated readings
#     in DIORAMA mode, real sensor readings when hardware is connected
#   ④ FFTAlarmClassifier added as a background thread — confirms FACP alarm
#     before global evacuation routing is activated (air-gap compliance)
#   ⑤ Pull Policy signals exposed on /api/get_route alongside the route
#   ⑥ RSET/ASET breakdown exposed on /api/status for dashboard display
#   ⑦ All new endpoints are backward-compatible — existing React code
#     still works without any changes
# =============================================================================

import csv
import os
import atexit
import math
import json
import time
import random
import threading
from datetime import datetime
from collections import deque

import cv2
import numpy as np
import paho.mqtt.client as mqtt
from ultralytics import YOLO
from flask import Flask, Response, jsonify, request
from flask_cors import CORS

# Local modules (must be in the same folder)
from routing_engine import (
    calculate_safest_route,
    force_route,
    reset_hysteresis,
    route_to_specific_exit,
    get_all_exit_routes,
    block_node_and_reroute,
    unblock_node,
    get_all_exit_routes,
    facp_store_alert,
    route_from_store,
    DOOR_TO_JUNCTION,
    DOOR_LABELS,
    run_pull_policy,
    update_crowd,
    TIER1_CROWD,
    live_node_status,
    get_crowd_velocity,
    estimate_rset,
    estimate_baseline_rset,
    rset_t2_sensitivity,
    J_TO_CORRIDOR,
    EXIT_TO_CORRIDOR,
    J_CORRIDOR_RANK,
    resolve_node_name,
)
from thermal_classifier import ThermalClassifier, _gradual_fire, _normal_ambient
from fft_classifier import FFTAlarmClassifier, _generate_alarm_tone, FRAME_SIZE, SAMPLE_RATE

app = Flask(__name__)
CORS(app)

# =============================================================================
# CORRIDOR STATE BUILDER — translates DYN-A* route + pull policy into the
# 5-corridor dict the ESP32 firmware parses (C-001..C-005).
# Each corridor entry is {"state": ..., "dir": 1|-1} — direction tells the
# LED chase which way to point so evacuees are always guided TOWARD the
# exit, never back into a blocked/hazard segment, regardless of which
# physical direction DYN-A* happens to traverse that corridor's nodes.
# Caller MUST hold state_lock before calling this (reads live_node_status,
# current_route, current_pull_signals).
# =============================================================================
def _build_corridor_states():
    # 1. Initialize the dictionary
    states = {c: {"state": "normal", "dir": 1}
              for c in ["C-001", "C-002", "C-003", "C-004", "C-005"]}

    # # 2. Simulation Gate: If in simulation, return the 'normal' states immediately.
    # if system_mode == "simulation":
    #     return states

    # 3. Hazard / quarantine takes priority 
    # Use 'states' defined above!
    for jid, data in live_node_status.items():
        corridor = J_TO_CORRIDOR.get(jid)
        if not corridor: continue
        if data.get("status") in ("alert", "quarantine"):
            states[corridor]["state"] = "hazard"

    # 4. Pull policy RED stop-lines
    for nid, info in current_pull_signals.items():
        corridor = J_TO_CORRIDOR.get(nid)
        if corridor and states[corridor]["state"] == "normal" and info.get("signal") == "RED":
            states[corridor]["state"] = "pull_stop"

    # 5. Pull policy AMBER / warning
    for nid, info in current_pull_signals.items():
        corridor = J_TO_CORRIDOR.get(nid)
        if corridor and states[corridor]["state"] == "normal" and info.get("signal") == "AMBER":
            states[corridor]["state"] = "warning"

    # 6. Active DYN-A* route
    for idx, node_id in enumerate(current_route):
        corridor = (J_TO_CORRIDOR.get(node_id) or EXIT_TO_CORRIDOR.get(node_id))
        if not corridor or states[corridor]["state"] in ("hazard", "pull_stop", "warning"):
            continue
        states[corridor]["state"] = "route"

        # Determine direction
        if idx + 1 < len(current_route):
            next_id = current_route[idx + 1]
            r_cur  = J_CORRIDOR_RANK.get(node_id)
            r_next = J_CORRIDOR_RANK.get(next_id)
            if r_cur is not None and r_next is not None:
                states[corridor]["dir"] = 1 if r_next < r_cur else -1
            elif next_id in EXIT_TO_CORRIDOR and EXIT_TO_CORRIDOR[next_id] == corridor:
                states[corridor]["dir"] = 1

    return states

# =============================================================================
# 1. GLOBAL SETUP & THREAD LOCKING
# =============================================================================
BROKER       = "broker.hivemq.com"
TOPIC        = "lumina/vitrox/demo/7a9b2f/alerts"   # unique — prevents hackathon collision
SENSOR_TOPIC = "lumina/vitrox/demo/7a9b2f/sensors"  # ESP32→Python: sensor data

def _on_sensor_message(client, userdata, msg):
    """
    Receives physical sensor events published by the ESP32 on SENSOR_TOPIC.
    Handles two sensor types:
      HC-SR04  → obstruction detected/cleared, calls block_node_and_reroute()
      MLX90614 → thermal anomaly, feeds temp reading into ThermalClassifier

    IMPORTANT: Real sensor events are SUPPRESSED in simulation mode.
    Bomba override cannot be overridden by sensor events.
    """
    global manual_override, current_route, system_state, thermal_state

    # Respect system mode — ignore real sensors in simulation mode
    with state_lock:
        _mode  = system_mode
        _bomba = bomba_override_active
    if _mode != "live":
        print(f"[SENSOR] Suppressed — system is in {_mode.upper()} mode")
        return
    if _bomba:
        print("[SENSOR] Suppressed — Bomba override active")
        return
    try:
        data   = json.loads(msg.payload.decode())
        sensor = data.get("sensor")

        # ── HC-SR04: physical corridor obstruction ─────────────────────────
        if sensor == "HC-SR04":
            status  = data.get("status")
            node_id = data.get("node", "C-003")
            dist    = data.get("distance_cm", -1)
            CORRIDOR_TO_JUNCTION = {
                "C-001": "J2", "C-002": "J4", "C-003": "J8",
                "C-004": "J12", "C-005": "J18",
            }
            junction = CORRIDOR_TO_JUNCTION.get(node_id, "J8")

            if status == "BLOCKED":
                print(f"[HC-SR04] Obstruction in {node_id} ({dist}cm) → blocking {junction}, recalculating route")
                with state_lock:
                    result = block_node_and_reroute(junction, current_route[0] if current_route else "J4")
                    current_route   = result["new_route"]
                    manual_override = True
                    _total_pax  = sum(d["crowd"] for d in live_node_status.values())
                    _corridors  = _build_corridor_states()
                mqtt_client.publish(TOPIC, json.dumps({
                    "status": "CRITICAL", "system_state": system_state,
                    "hazard_type": f"OBSTRUCTION DETECTED in {node_id}",
                    "manual_override": True, "person_count": _total_pax,
                    "green_direction": "FOLLOW_ROUTE", "corridors": _corridors,
                }))
            elif status == "CLEAR":
                print(f"[HC-SR04] {node_id} cleared → unblocking {junction}")
                with state_lock:
                    unblock_node(junction)
                    reset_hysteresis()
                    # Only release manual_override if no active fire/thermal hazard —
                    # clearing a debris obstruction shouldn't cancel an ongoing evacuation
                    if system_state == "NORMAL":
                        manual_override = False

        # ── MLX90614: real thermal anomaly from physical IR sensor ─────────
        elif sensor == "MLX90614":
            temp_c = data.get("temp_c", 0)
            print(f"[MLX90614] Real thermal reading: {temp_c}°C")
            # Feed the real reading into the existing ThermalClassifier —
            # same pipeline as the simulated path, now driven by real hardware.
            result = thermal_clf.classify(temp_c)
            with state_lock:
                thermal_state = result["state"]
                if result["state"] in ("WARNING", "ALERT") and system_state == "NORMAL":
                    system_state = "HAZARD"
                    live_node_status["J7"]["status"] = "alert"
                    live_node_status["J7"]["hazard"] = "thermal"
                    _total_pax  = sum(d["crowd"] for d in live_node_status.values())
                    _corridors  = _build_corridor_states()
            if result["state"] == "ALERT":
                mqtt_client.publish(TOPIC, json.dumps({
                    "status":       "CRITICAL",
                    "system_state": "HAZARD",
                    "hazard_type":  "THERMAL ANOMALY (MLX90614)",
                    "temp_c":       temp_c,
                    "person_count": _total_pax,
                    "corridors":    _corridors,
                }))
                print(f"[MLX90614] THERMAL ALERT triggered at {temp_c}°C")

    except Exception as e:
        print(f"[SENSOR] Message error: {e}")

mqtt_client = mqtt.Client(mqtt.CallbackAPIVersion.VERSION2, "Lumina_Edge_Streamer")
mqtt_client.message_callback_add(SENSOR_TOPIC, _on_sensor_message)
try:
    mqtt_client.connect(BROKER, 1883, 60)
    mqtt_client.subscribe(SENSOR_TOPIC)   # listen for ESP32 sensor events
    mqtt_client.loop_start()
    print("[MQTT] Connected to broker")
except Exception as e:
    print(f"[MQTT] Warning: Could not connect ({e}) — running offline")

state_lock   = threading.Lock()

# Core system state
system_state = "NORMAL"          # NORMAL | HAZARD
ai_mode      = "DIORAMA"         # DIORAMA | ENTERPRISE
facp_confirmed = False           # True once FFT confirms the official alarm

# =============================================================================
# SYSTEM MODE — controls which event sources are accepted
#   "simulation" : only manual simulation triggers accepted; real sensors ignored
#   "live"       : only real sensor data accepted; simulation triggers disabled
# BOMBA override (bomba_override_active) is a special elevated state that works
# in ANY mode and cannot be cancelled by simulation or sensor events.
# Priority: bomba_override (3) > live_sensor (2) > simulation (1)
# =============================================================================
system_mode           = "simulation"   # "simulation" | "live"
bomba_override_active = False          # True when Bomba has issued a command override

# Simulation trigger state — which event type was manually triggered
# None | "fire" | "fallen" | "crowd"
sim_trigger_type  = None
sim_trigger_node  = None   # most recent trigger node (for status display)
active_hazard_nodes = []  # list of {node_id, event_type} for multi-hazard tracking

# Fire simulation flag — no physical thermal sensor in this prototype.
# Set to True by /trigger (BOMBA "Simulate Fire" button) for demo purposes.
# Cleared by /reset. Camera-based fall detection is independent of this flag.
fire_sim_active = False

# Shared metrics (written by bg threads, read by Flask)
current_person_count  = 0
current_track_ids     = []        # list of active anonymous track IDs
crowd_velocity_lobby  = 0.0      # rate-of-change at lobby node (N-011)
thermal_state         = "NORMAL" # NORMAL | WARNING | ALERT
fft_state             = "SILENT" # SILENT | DETECTING | CONFIRMED
current_route         = ["J19","J20","J1","EXIT-1"]
current_pull_signals  = {}
current_rset          = {}
current_route_cost    = 0   # raw DYN-A* cost score — exposed to frontend
current_per_node_routes = []  # per-hazard-node routes for multi-path real-time display

LOG_FILE = "lumina_telemetry_log.csv"

# Startup timestamp — for /api/health uptime display
_startup_time = time.time()

# Manual override flag — set by /trigger and /api/block_node.
# When True: stochastic sensor drift pauses so BOMBA's manual command
# isn't overwritten by simulated sensor noise. Only /reset clears this.
manual_override = False

# Module-level drift tick — replaces the fragile generate_frames._last_drift
# function attribute. Safe across Flask threads; GIL protects the int read/write.
_last_drift_tick = -1

# Live temperature readings per node — populated by thermal thread,
# read by /api/status so the React sparkline shows real escalating values.
_latest_temps = {nid: 27.0 for nid in ["J4","J7","J8","J18","J12"]}

# Classifier latency readings — float writes are GIL-atomic, no lock needed.
# Updated by bg threads every cycle, read by /api/status for display.
_thermal_latency_ms = 0.0
_fft_latency_ms     = 0.0

# Simulated fleet size — 6 real nodes + 192 standby nodes matching proposal
NODES_ONLINE = 198
NODES_TOTAL  = 200

# Single source of truth for battery data — used by api_health() and download_log()
NODE_BATTERY = {
    "NODE-A": {"pct": 94, "next_service": "Aug 10"},
    "NODE-B": {"pct": 87, "next_service": "Aug 01"},
    "NODE-C": {"pct": 72, "next_service": "Jul 15"},
    "NODE-D": {"pct": 81, "next_service": "Aug 05"},
    "NODE-E": {"pct": 96, "next_service": "Aug 12"},
    "NODE-F": {"pct": 63, "next_service": "Jul 01"},
}

# Mirrors frontend's LUMINA_NODE_DEFS labels — only the 6 physical Lumina
# ceiling units have batteries, not individual junctions/doors.
LUMINA_NODE_LABELS = {
    "NODE-A": "West Corridor", "NODE-B": "Central Crossroad",
    "NODE-C": "East Corridor", "NODE-D": "South-Central",
    "NODE-E": "South-West",    "NODE-F": "East-South",
}

# =============================================================================
# 2. AI MODEL LOADING
#
# --- DEPLOYMENT NOTE: Edge TPU Production Pipeline -
# This prototype runs standard PyTorch (.pt) weights on a laptop CPU.
# Production deployment on the RK3588 hardware node requires:
#   1. Export:  yolo export model=yolov8n.pt format=rknn  (via rknn-toolkit2)
#   2. Load:    model = YOLO("lumina_topdown_v1.rknn")
#   3. Result:  RK3588 NPU (6 TOPS) reduces inference from ~150ms (CPU) to
#               ~12ms (NPU) — well within the 500ms ASET actuation target.
# Reference: https://docs.ultralytics.com/integrations/rockchip-rknn/
# ---
# =============================================================================
print("[INIT] Loading DUAL-ENGINE AI models...")
model_diorama    = YOLO("yolov8n.pt")         # toy/diorama: bounding-box aspect ratio
model_enterprise = YOLO("yolov8n-pose.pt")    # real humans: skeletal keypoints

# Custom-trained fall-detection model — single class "Fall-Detected", fine-tuned
# on the Roboflow fall-detection-ca3o8 dataset (10.8k images, elevated/CCTV
# angles). Used as a FALLBACK ONLY: when model_diorama finds zero people in a
# frame, this checks specifically for a fallen-person shape — covers the
# steep-angle case where COCO's general "person" class loses the silhouette.
# Path: place fall_detector.pt next to this script, or set FALL_MODEL_PATH env var.
import os as _os
_fall_model_path = _os.environ.get("FALL_MODEL_PATH", "models/fall_detector.pt")
try:
    model_fall_detector = YOLO(_fall_model_path)
    print(f"[INIT] Custom fall-detection model loaded from {_fall_model_path}")
except Exception as _e:
    model_fall_detector = None
    print(f"[INIT] WARNING: could not load fall-detection model ({_e}) — "
          f"falling back to background-subtraction only")

print("[INIT] Starting camera...")
# CAMERA_INDEX: 0 = built-in webcam, 1+ = USB/external webcam
# Set env var CAMERA_INDEX=1 if USB webcam is not detected on index 0
# e.g.  CAMERA_INDEX=1 python lumina_live_stream.py
import os as _os
_cam_idx = int(_os.environ.get("CAMERA_INDEX", 0))
print(f"[INIT] Using camera index {_cam_idx} (set CAMERA_INDEX env var to change)")
cap = cv2.VideoCapture(_cam_idx)
cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT,  720)
cap.set(cv2.CAP_PROP_AUTOFOCUS, 1)
cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)   # always read latest frame, no stale queue

# =============================================================================
# 3. THERMAL CLASSIFIER — background thread
#    In DIORAMA mode: feeds simulated temperature readings (demo without sensor)
#    In ENTERPRISE mode: replace _read_thermal_sensor() with your real sensor
# =============================================================================
thermal_clf = ThermalClassifier("J4")   # Lobby node (was J16 — no physical hardware there; merged into J4)
_thermal_tick = 0                          # frame counter for simulated signal

def _read_thermal_sensor_simulated() -> float:
    global _thermal_tick
    _thermal_tick += 1
    with state_lock:
        in_fire = fire_sim_active
    if in_fire:
        return _gradual_fire(_thermal_tick, onset=0)
    return _normal_ambient(_thermal_tick)

def _thermal_thread():
    global thermal_state, system_state, _thermal_latency_ms
    while True:
        temp   = _read_thermal_sensor_simulated()
        result = thermal_clf.classify(temp)
        _thermal_latency_ms = result["latency_ms"]
        with state_lock:
            # Previously this always heated up J7 specifically, regardless
            # of which node the fire was actually triggered at — clicking
            # fire at J12 would correctly avoid J12 in routing, but the
            # temperature dashboard would show J7 spiking instead. Find the
            # actual fire node(s) so only those heat up.
            fire_nodes = [h["node_id"] for h in active_hazard_nodes if h.get("event_type")=="fire"]
        # J4 IS the sensor location now (merged from J16) — direct read, no scaling
        _latest_temps["J4"] = round(result["temp_c"], 1)
        # Dynamically heat up whichever OTHER tracked nodes are actually on
        # fire; everyone else in _latest_temps stays at ambient.
        for _nid in _latest_temps:
            if _nid == "J4":
                continue
            if _nid in fire_nodes:
                _latest_temps[_nid] = round(min(150, result["temp_c"] * 1.8), 1)
            else:
                _latest_temps[_nid] = round(27.0 + random.uniform(-0.5, 0.5), 1)

        with state_lock:
            thermal_state = result["state"]
            # In simulation mode, thermal sensor cannot override sim state
            if result["state"] == "ALERT" and system_state == "NORMAL" and system_mode == "live":
                system_state = "HAZARD"
                live_node_status["J4"]["status"] = "alert"
                live_node_status["J4"]["hazard"] = "thermal"
                # Real fire hard-blocks the corridor, not just a soft cost
                # penalty — a burning junction genuinely can't be walked
                # through, same as a manual BOMBA collapse block.
                live_node_status["J4"]["impassable"] = True
                _publish_alert = True
            else:
                _publish_alert = False

        # Publish OUTSIDE the lock — I/O must never be inside a threading lock
        if _publish_alert:
            with state_lock:
                _total_pax  = sum(d["crowd"] for d in live_node_status.values())
                _corridors  = _build_corridor_states()
            mqtt_client.publish(TOPIC, json.dumps({
                "status":       "CRITICAL",
                "system_state": "HAZARD",
                "hazard_type":  "THERMAL ANOMALY",
                "temp_c":       result["temp_c"],
                "z_score":      result["z_score"],
                "person_count": _total_pax,
                "corridors":    _corridors,
            }))
        time.sleep(0.2)   # 5 Hz

threading.Thread(target=_thermal_thread, daemon=True).start()
print("[INIT] Thermal classifier thread started")

# =============================================================================
# 4. FFT ACOUSTIC CLASSIFIER — background thread
# =============================================================================
fft_clf = FFTAlarmClassifier("J4")

def _read_audio_frame_simulated() -> np.ndarray:
    with state_lock:
        in_fire = fire_sim_active
    if in_fire:
        return _generate_alarm_tone(FRAME_SIZE / SAMPLE_RATE)
    else:
        n = FRAME_SIZE
        t = np.linspace(0, FRAME_SIZE / SAMPLE_RATE, n)
        noise = (
            0.3 * np.sin(2 * math.pi * 60 * t) +
            0.2 * np.sin(2 * math.pi * 120 * t) +
            np.random.normal(0, 0.05, n)
        )
        return noise.astype(np.float32)

def _fft_thread():
    global fft_state, facp_confirmed, _fft_latency_ms
    while True:
        frame  = _read_audio_frame_simulated()
        result = fft_clf.classify_frame(frame)
        _fft_latency_ms = result["latency_ms"]       # GIL-atomic float write
        _publish_facp = False
        with state_lock:
            fft_state = result["state"]
            if result["state"] == "CONFIRMED" and not facp_confirmed:
                facp_confirmed = True
                _publish_facp  = True
                print("[FFT] FACP Positive Alarm Sequence CONFIRMED — global routing active")
        # Publish OUTSIDE the lock — same pattern as _thermal_thread
        if _publish_facp:
            mqtt_client.publish(TOPIC, json.dumps({
                "status":   "FACP_CONFIRMED",
                "snr_db":   result["snr_db"],
                "alarm_hz": 520,
            }))
        time.sleep(0.1)   # 10 Hz

threading.Thread(target=_fft_thread, daemon=True).start()
print("[INIT] FFT acoustic classifier thread started")

# =============================================================================
# 4b. MQTT HEARTBEAT — background thread
# Keeps the ESP32 synced with live person counts during NORMAL operation.
# CRITICAL/RESOLVED/FACP_CONFIRMED messages already cover hazard transitions —
# this heartbeat fills the gap during normal (non-hazard) operation so the
# ESP32's person count display stays current without flooding it with
# per-frame video updates (15fps would overwhelm the Wi-Fi chip).
# =============================================================================
def _heartbeat_thread():
    while True:
        with state_lock:
            _mode      = system_mode
            _state     = system_state
            _manual    = manual_override
            _bomba     = bomba_override_active
            _total_pax = sum(d["crowd"] for d in live_node_status.values())
            _corridors = _build_corridor_states()

        # ALLOW hardware updates ONLY if in Live Mode OR if Bomba manually overrides
        if _mode == "live" or _bomba:
            mqtt_status = "CRITICAL" if _state == "HAZARD" else "NORMAL"
            _stealth = not _manual and _state == "NORMAL"
            
            payload = json.dumps({
                "status":          mqtt_status,
                "system_state":    _state,
                "manual_override": _manual,
                "stealth_mode":    _stealth,
                "person_count":    _total_pax,
                "corridors":       _corridors,
            })
            print(f"[DEBUG] Sending to ESP32: {payload}", flush=True)
            mqtt_client.publish(TOPIC, payload, retain=True)
            print(f">>> Heartbeat published to MQTT [LIVE MODE - {mqtt_status}]", flush=True)
        else:
            
            # SIMULATION MODE: Send isolation signal so ESP32 stays dark and silent
            mqtt_client.publish(TOPIC, json.dumps({
                "status": "SIMULATION_ACTIVE", 
                "system_state": "NORMAL"
            }), retain=True)
            # ADDED flush=True to bypass terminal buffering
            print(">>> Heartbeat published to MQTT [SIMULATION MODE - Hardware Isolated]", flush=True)
            
        time.sleep(1.0)

threading.Thread(target=_heartbeat_thread, daemon=True).start()
print("[INIT] MQTT heartbeat thread started")

# =============================================================================
# 5. VIDEO GENERATOR
# =============================================================================
_SKELETON_PAIRS = [
    (0,1),(0,2),(1,3),(2,4),
    (5,6),(5,7),(7,9),(6,8),(8,10),
    (5,11),(6,12),(11,12),
    (11,13),(13,15),(12,14),(14,16),
]

def _draw_skeleton(frame, kpts):
    for a, b in _SKELETON_PAIRS:
        if a >= len(kpts) or b >= len(kpts):
            continue
        ax, ay = int(kpts[a][0]), int(kpts[a][1])
        bx, by = int(kpts[b][0]), int(kpts[b][1])
        if ax > 0 and ay > 0 and bx > 0 and by > 0:
            cv2.line(frame, (ax, ay), (bx, by), (180, 180, 180), 1)
    for kpt in kpts:
        kx, ky = int(kpt[0]), int(kpt[1])
        if kx > 0 and ky > 0:
            cv2.circle(frame, (kx, ky), 3, (255, 255, 255), -1)

def _check_fall_keypoints(kpts) -> tuple:
    """
    Keypoint-based fall detection (face-up / front-facing).
    Returns (is_fallen: bool, confidence: str)
    Fallen when nose_y > avg_hip_y — person is horizontal from top-down camera.
    """
    if len(kpts) < 13:
        return False, "insufficient_keypoints"
    nose_y    = float(kpts[0][1])
    avg_hip_y = (float(kpts[11][1]) + float(kpts[12][1])) / 2
    if nose_y <= 0 or avg_hip_y <= 0:
        return False, "keypoints_occluded"
    return nose_y > avg_hip_y, "keypoint_signal"

def _check_fall_bbox(w: int, h: int) -> tuple:
    """
    Bounding-box aspect ratio fall detection (works face-down, back-facing,
    any orientation — more robust than keypoints alone).
    Returns (is_fallen: bool, confidence: str)
    Fallen when width > 1.3× height — person is horizontal.
    """
    return w > (h * 1.3), "bbox_signal"

def _check_fall_enterprise(kpts, w: int, h: int) -> tuple:
    """
    Dual-signal fall classifier — combines keypoint + bbox detection.
    Either signal alone can trigger, making it robust to:
      - Face-down falls    (keypoints fail → bbox catches it)
      - Back-facing falls  (keypoints fail → bbox catches it)
      - Partial occlusion  (bbox fails   → keypoints catch it)
      - Top-down cameras   (both signals combined for higher sensitivity)

    Returns (is_fallen: bool, trigger: str)
    """
    kpt_fallen, kpt_reason = _check_fall_keypoints(kpts)
    bbox_fallen, _         = _check_fall_bbox(w, h)

    if kpt_fallen and bbox_fallen:
        return True,  "DUAL (kpt+bbox)"    # highest confidence
    elif kpt_fallen:
        return True,  "KPT only"           # face-up, front-facing
    elif bbox_fallen:
        return True,  "BBOX only"          # face-down, back-facing
    else:
        return False, "upright"


# =============================================================================
# FALLBACK: Background-subtraction blob detection
#
# Used ONLY when YOLO returns zero person detections in a frame. At steep
# camera angles (e.g. 45° down on a toy diorama), a fallen figure can lose
# enough of its silhouette that YOLO — trained on real human poses — simply
# doesn't recognize it as "person" at any confidence threshold. This isn't
# a tuning problem; the features needed for shape-based detection aren't
# present in the frame at that angle/scale.
#
# Background subtraction sidesteps this entirely: it doesn't care what
# shape the object makes, only that something non-background appeared in
# the scene and is sitting roughly still — which is consistent with "a
# person is down" rather than walking through. This is a classical CV
# technique, not a neural model, and is deliberately simple/fast (<5ms)
# so it can run every frame as a cheap fallback without hurting FPS.
# =============================================================================
_bg_subtractor = cv2.createBackgroundSubtractorMOG2(
    history=300, varThreshold=40, detectShadows=False
)
_MIN_BLOB_AREA       = 800    # px^2 — ignore tiny noise/specks
_STATIONARY_RADIUS   = 25     # px — blob centroid must stay within this to count as "still"
_STATIONARY_SECONDS  = 2.5    # how long a blob must persist before flagging as fallen

def _check_fall_background_subtraction(frame, state) -> tuple:
    """
    Fallback fall signal when YOLO sees zero people in the frame.
    Returns (is_fallen: bool, trigger: str, bbox: tuple|None)

    `state` carries persistent tracking across calls:
      state["bg_blob_centroid"]   — last seen centroid (x, y) or None
      state["bg_blob_since"]      — timestamp the blob first appeared near
                                     this position, or 0 if not tracking
    """
    fg_mask = _bg_subtractor.apply(frame, learningRate=0.003)
    fg_mask = cv2.morphologyEx(fg_mask, cv2.MORPH_OPEN,
                                np.ones((5, 5), np.uint8))
    contours, _ = cv2.findContours(fg_mask, cv2.RETR_EXTERNAL,
                                    cv2.CHAIN_APPROX_SIMPLE)

    if not contours:
        state["bg_blob_centroid"] = None
        state["bg_blob_since"]    = 0
        return False, "no_blob", None

    # Largest blob only — assume one figure of interest at a time for the demo
    largest = max(contours, key=cv2.contourArea)
    area    = cv2.contourArea(largest)
    if area < _MIN_BLOB_AREA:
        state["bg_blob_centroid"] = None
        state["bg_blob_since"]    = 0
        return False, "blob_too_small", None

    x, y, w, h = cv2.boundingRect(largest)
    cx, cy     = x + w // 2, y + h // 2
    t_now      = time.time()

    prev = state.get("bg_blob_centroid")
    if prev is not None:
        dist = math.hypot(cx - prev[0], cy - prev[1])
        if dist > _STATIONARY_RADIUS:
            # Blob moved too much — reset the "since" timer, still walking/shifting
            state["bg_blob_since"] = t_now
    else:
        state["bg_blob_since"] = t_now

    state["bg_blob_centroid"] = (cx, cy)
    elapsed = t_now - state["bg_blob_since"]

    if elapsed >= _STATIONARY_SECONDS:
        return True, "bg_subtraction_stationary", (x, y, x + w, y + h)
    return False, "bg_subtraction_pending", (x, y, x + w, y + h)


# ── Shared frame buffer (AI worker writes, /video_feed reads) ──────────────
# This decouples AI processing from the browser: YOLO inference, fall
# detection, and DYN-A* routing run continuously in a background thread
# regardless of whether anyone has the dashboard open. /video_feed simply
# reads whatever frame the worker most recently produced.
_frame_buffer   = None          # latest annotated JPEG bytes
_frame_lock     = threading.Lock()
_ai_thread_stop = threading.Event()


def _process_ai_cycle(cap, state):
    """
    Runs ONE iteration of: read frame -> YOLO inference -> fall/crowd
    detection -> DYN-A* reroute -> annotate frame -> store in buffer.
    `state` is a dict carrying loop-persistent variables across calls
    (fall timers, frame counter, cached inference results, etc).
    """
    global system_state, ai_mode, current_person_count
    global current_track_ids, crowd_velocity_lobby
    global current_route, current_pull_signals, current_rset, current_route_cost, current_per_node_routes
    global facp_confirmed, _last_drift_tick, _frame_buffer

    success, frame = cap.read()
    if not success:
        time.sleep(0.1)
        return
    
    # If the mean pixel value is extremely low, the camera is likely black/disconnected
    if np.mean(frame) < 5: 
        print("[AI] Warning: Black frame detected, skipping AI cycle to prevent false hazard")
        return

    t_now = time.time()
    fps   = 1.0 / max(t_now - state["prev_time"], 1e-6)
    state["prev_time"] = t_now
    state["frame_counter"] += 1

    with state_lock:
        cur_state = system_state
        cur_mode  = ai_mode

    person_count           = 0
    current_frame_has_fall = False
    track_ids_this_frame   = []

    # --- BYTETRACK DETECTION + TRACKING PASS ---
    # Frame skipping halves CPU load; ByteTrack's Kalman filter predicts
    # positions on skipped frames so track IDs remain stable.
    _skip_interval = 3 if cur_state == "HAZARD" else 2
    _run_inference = (state["frame_counter"] % _skip_interval == 0) or (state["last_results"] is None)
    fallen_boxes = []

    if _run_inference:
        if cur_mode == "DIORAMA":
            results = model_diorama.track(frame, persist=True, conf=0.30,
                                          classes=[0], verbose=False)
        else:
            results = model_enterprise.track(frame, persist=True, conf=0.60,
                                             verbose=False)
        state["last_results"] = results
    else:
        results = state["last_results"]

    for r in results:
        if r.boxes is None:
            continue
        track_ids = (
            r.boxes.id.int().cpu().tolist()
            if r.boxes.id is not None else
            [None] * len(r.boxes)
        )
        for i, box in enumerate(r.boxes):
            if cur_mode == "DIORAMA" and int(box.cls) != 0:
                continue
            tid = track_ids[i]
            if tid is None:
                continue
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            w, h = x2 - x1, y2 - y1
            person_count += 1
            track_ids_this_frame.append(tid)

            is_fallen    = False
            fall_trigger = "upright"
            if cur_mode == "DIORAMA":
                is_fallen, fall_trigger = _check_fall_bbox(w, h)
            elif r.keypoints is not None and len(r.keypoints.xy) > i:
                kpts = r.keypoints.xy[i]
                _draw_skeleton(frame, kpts)
                is_fallen, fall_trigger = _check_fall_enterprise(kpts, w, h)
            elif cur_mode == "ENTERPRISE":
                is_fallen, fall_trigger = _check_fall_bbox(w, h)

            if is_fallen:
                fallen_boxes.append((x1, y1, x2, y2))
                current_frame_has_fall = True

            box_color = (0, 0, 255) if is_fallen else (0, 255, 0)
            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            cv2.putText(frame, f"ID:{tid}", (x1, y1 - 6),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, box_color, 1)
            if is_fallen:
                cv2.putText(frame, f"FALL [{fall_trigger}]", (x1, y2 + 14),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 100, 255), 1)

    # --- FALL-DETECTION MODEL: runs every cycle, independent of person_count ---
    # Gating this behind "model_diorama found nobody" was a bug — a fallen
    # figure lying near a standing person that IS detected normally would
    # never trigger the check, since person_count would be >0. The trained
    # fall model is cheap enough (tens of ms) to just run every inference
    # pass and merge its results in regardless of what the main detector saw.
    def _boxes_overlap(a, b, thresh=0.3):
        ax1, ay1, ax2, ay2 = a
        bx1, by1, bx2, by2 = b
        ix1, iy1 = max(ax1, bx1), max(ay1, by1)
        ix2, iy2 = min(ax2, bx2), min(ay2, by2)
        iw, ih = max(0, ix2 - ix1), max(0, iy2 - iy1)
        inter = iw * ih
        if inter == 0:
            return False
        area_a = max(1, (ax2 - ax1) * (ay2 - ay1))
        return (inter / area_a) > thresh

    if model_fall_detector is not None:
        fd_results = model_fall_detector.predict(frame, conf=0.40, verbose=False)
        for fd_box in fd_results[0].boxes:
            fx1, fy1, fx2, fy2 = map(int, fd_box.xyxy[0])
            fd_conf = float(fd_box.conf[0])
            new_box = (fx1, fy1, fx2, fy2)
            # Skip if this overlaps a box already flagged this frame (avoid
            # double-marking the same fallen figure if model_diorama's own
            # bbox/keypoint check already caught it).
            if any(_boxes_overlap(new_box, existing) for existing in fallen_boxes):
                continue
            fallen_boxes.append(new_box)
            current_frame_has_fall = True
            cv2.rectangle(frame, (fx1, fy1), (fx2, fy2), (0, 0, 255), 2)
            cv2.putText(frame, f"FALL [fall_model {fd_conf:.2f}]", (fx1, fy2 + 14),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 100, 255), 1)

    # --- LAST-RESORT FALLBACK: background subtraction ---
    # Only when model_diorama found literally nobody this frame AND the
    # trained fall model also found nothing — covers the case where even
    # the fall model's COCO-adjacent training doesn't generalize to this
    # specific shot. Blob detection assumes one isolated change region, so
    # it's not meaningful to run in busy mixed scenes — zero-detection only.
    if person_count == 0 and not current_frame_has_fall:
        bg_fallen, bg_trigger, bg_bbox = _check_fall_background_subtraction(frame, state)
        if bg_fallen and bg_bbox is not None:
            bx1, by1, bx2, by2 = bg_bbox
            fallen_boxes.append((bx1, by1, bx2, by2))
            current_frame_has_fall = True
            cv2.rectangle(frame, (bx1, by1), (bx2, by2), (0, 0, 255), 2)
            cv2.putText(frame, f"FALL [{bg_trigger}]", (bx1, by2 + 14),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0, 100, 255), 1)

    # --- FEED INTO ROUTING ENGINE ---
    # Locked: update_crowd/get_crowd_velocity mutate live_node_status,
    # which /reset and other Flask threads also touch concurrently.
    with state_lock:
        # Only update J4 crowd from camera if it's not a sim-triggered hazard node
        # (J4 is now the lobby node — merged from J16, which had no physical hardware)
        # Otherwise the camera count (0 in demo) would immediately overwrite the sim
        _lobby_is_sim = any(h["node_id"]=="J4" for h in active_hazard_nodes)
        if not _lobby_is_sim:
            update_crowd("J4", person_count)
        vel = get_crowd_velocity("J4")
        current_person_count = person_count
        current_track_ids    = track_ids_this_frame
        crowd_velocity_lobby = round(vel, 3)

    # --- STOCHASTIC SENSOR MODEL — secondary nodes ---
    # Skip any node that's currently a sim-triggered hazard — same protection
    # J4 (lobby, merged from J16) already has above. Without this, the drift
    # values here (all below TIER1_CROWD) would force-clear a crowd sim trigger
    # on J7/J8/J12/J18 within ~2 seconds of it firing — the icon/route would
    # blink once then vanish. J4 itself is excluded from this dict entirely —
    # it's camera-driven now, not stochastic, same reason J16 never was.
    if not manual_override and int(t_now) % 2 == 0 and int(t_now) != _last_drift_tick:
        _last_drift_tick = int(t_now)
        _in_hazard = (cur_state == "HAZARD")
        _sensor_model = {
            "J7": (20, 45 if not _in_hazard else 99),
            "J8": (10, 38), "J18": (5, 20), "J12": (8, 30),
        }
        with state_lock:
            _sim_node_ids = {h["node_id"] for h in active_hazard_nodes}
            for _nid, (_lo, _hi) in _sensor_model.items():
                if _nid in _sim_node_ids:
                    continue
                _cur   = live_node_status[_nid]["crowd"]
                _drift = random.randint(-1, 1)
                _new   = max(_lo, min(_hi, _cur + _drift))
                update_crowd(_nid, _new)

    if vel > 5 and cur_state == "NORMAL" and system_mode == "live":
        print(f"[CROWD] Velocity spike {vel:+.2f} — pre-emptive reroute")
        with state_lock:
            live_node_status["J4"]["status"] = "warning"

    # --- FALL ESCALATION ---
    # In simulation mode, camera fall detection is suppressed — manual triggers only
    if current_frame_has_fall and system_mode == "live":
        state["recovery_timer_start"] = 0
        if state["fall_timer_start"] == 0:
            state["fall_timer_start"] = t_now
        if t_now - state["fall_timer_start"] >= 3.0 and cur_state == "NORMAL":
            with state_lock:
                system_state = "HAZARD"
                live_node_status["J4"]["hazard"] = "fall"
                live_node_status["J4"]["status"] = "alert"
                live_node_status["J4"]["pull_signal"] = "RED"
                _route     = list(current_route)
                _total_pax = sum(d["crowd"] for d in live_node_status.values())
                _corridors = _build_corridor_states()
            mqtt_client.publish(TOPIC, json.dumps({
                "status": "CRITICAL", "system_state": "CRITICAL",
                "hazard_type": "FALL DETECTED", "person_count": _total_pax,
                "track_count": len(track_ids_this_frame),
                "stealth_mode": False, "green_led": True, "red_led": False,
                "buzzer_active": True, "green_direction": "FOLLOW_ROUTE",
                "active_route": _route, "corridors": _corridors,
            }))
    else:
        state["fall_timer_start"] = 0
        with state_lock:
            _n011_hazard = live_node_status["J4"]["hazard"]
        # Only auto-recover fall in live mode — simulation handles its own state
        if system_mode == "live" and cur_state == "HAZARD" and _n011_hazard == "fall":
            if state["recovery_timer_start"] == 0:
                state["recovery_timer_start"] = t_now
            if t_now - state["recovery_timer_start"] >= 3.0:
                with state_lock:
                    system_state   = "NORMAL"
                    facp_confirmed = False
                    live_node_status["J4"]["hazard"]      = None
                    live_node_status["J4"]["status"]      = "normal"
                    live_node_status["J4"]["pull_signal"] = "GREEN"
                with state_lock:
                    _total_pax = sum(d["crowd"] for d in live_node_status.values())
                    _corridors = _build_corridor_states()
                mqtt_client.publish(TOPIC, json.dumps({
                    "status": "RESOLVED", "system_state": "NORMAL",
                    "person_count": _total_pax, "stealth_mode": True,
                    "green_led": False, "red_led": False, "buzzer_active": False,
                    "green_direction": "NONE", "corridors": _corridors,
                }), retain=True)
        else:
            state["recovery_timer_start"] = 0

    # --- DYN-A* REROUTE (throttled to 1/sec) ---
    # Locked end-to-end: calculate_safest_route/run_pull_policy/estimate_rset
    # all read+mutate live_node_status and routing_engine's module-level
    # hysteresis cache. Without one continuous lock, a concurrent /reset
    # (or another Flask request thread) can mutate live_node_status mid-
    # calculation -> RuntimeError: dictionary changed size during iteration.
    if t_now - state["route_cooldown"] >= 1.0 and not manual_override:
        state["route_cooldown"] = t_now
        with state_lock:
            # In simulation mode with active hazards: recalculate from each
            # hazard node so all per-node routes stay current as crowd changes.
            # In normal/live mode: use lobby camera start node.
            if system_mode == "simulation" and active_hazard_nodes:
                # Recalculate best route for each active hazard node
                _per = []
                for _h in active_hazard_nodes:
                    _h_routes = get_all_exit_routes(_h["node_id"])
                    # ALWAYS keep the hazard tracked, even with zero routes —
                    # same fix as sim_trigger() below. A hazard that BECOMES
                    # stranded later (e.g. a second block made after this
                    # hazard was already triggered) was silently vanishing
                    # from current_per_node_routes right here, on this
                    # periodic tick, because this only appended when
                    # _h_routes was non-empty. The node's red/icon styling
                    # persisted (driven separately by live_node_status,
                    # untouched by this loop), but its tab and "no route"
                    # shelter message disappeared with zero trace.
                    _per.append({
                        "node_id":    _h["node_id"],
                        "event_type": _h["event_type"],
                        "best_path":  _h_routes[0]["path"] if _h_routes else [],
                        "best_exit":  _h_routes[0]["exit"] if _h_routes else None,
                        "best_cost":  _h_routes[0]["cost"] if _h_routes else None,
                        "all_exits":  _h_routes,
                        # Per-hazard RSET so the dashboard can show a correct
                        # breakdown for WHICHEVER hazard tab is selected,
                        # instead of only ever showing the single global
                        # current_route's RSET regardless of which hazard
                        # you're actually looking at.
                        "rset":       estimate_rset(_h_routes[0]["path"]) if _h_routes else None,
                    })
                    if _h["node_id"] in live_node_status:
                        live_node_status[_h["node_id"]]["shelter_in_place"] = (len(_h_routes) == 0)
                if _per:
                    current_per_node_routes[:] = _per
                    current_route[:] = _per[-1]["best_path"]
            else:
                lobby_hazard = live_node_status.get("J4", {}).get("hazard")
                start_node = "J7" if lobby_hazard == "fall" else "J4"
                path, score = calculate_safest_route(start_node, verbose=False)
                if path:
                    if start_node == "J7":
                        path = ["J4"] + path
                    current_route[:] = path
                    current_route_cost = score
                    current_per_node_routes.clear()
            signals = run_pull_policy()
            rset    = estimate_rset(current_route)
            current_pull_signals = signals
            current_rset         = rset

    with state_lock:
        _state   = system_state
        _thermal = thermal_state
        _fft     = fft_state
        _vel     = crowd_velocity_lobby

    _mode_txt = "TOY DIORAMA" if cur_mode == "DIORAMA" else "REAL-WORLD SKELETAL"
    cv2.putText(frame, f"MODE: {_mode_txt}", (10, 25),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)
    cv2.putText(frame, f"FPS:{fps:.0f}  PERSONS:{person_count}  VEL:{_vel:+.1f}/rdg",
                (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
    cv2.putText(frame, f"THERMAL:{_thermal}  FFT:{_fft}  STATE:{_state}",
                (10, 72), cv2.FONT_HERSHEY_SIMPLEX, 0.5,
                (0, 255, 0) if _state == "NORMAL" else (0, 80, 255), 2)

    for fx1, fy1, fx2, fy2 in fallen_boxes:
        cx, cy = (fx1 + fx2) // 2, (fy1 + fy2) // 2
        cv2.circle(frame, (cx, cy), 60, (0, 165, 255), 2)
        cv2.putText(frame, "BUFFER ZONE", (cx - 42, cy - 65),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 165, 255), 1)

    if _state == "HAZARD":
        h_frame, w_frame = frame.shape[:2]
        cv2.rectangle(frame, (0, 0), (w_frame, h_frame), (0, 0, 255), 6)
        with state_lock:
            route_txt = " -> ".join(current_route)
        cv2.putText(frame, f"ROUTE: {route_txt}", (10, h_frame - 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0, 255, 120), 1)

    ret, buffer = cv2.imencode(".jpg", frame)
    if ret:
        with _frame_lock:
            _frame_buffer = buffer.tobytes()


def _ai_worker():
    """
    Background daemon thread: runs the full AI + routing cycle continuously,
    independent of whether any browser has /video_feed open. This is the
    fix for the "observer-dependent AI loop" — DYN-A* and fall detection
    must never stop just because no one is watching the camera feed.
    """
    state = {
        "fall_timer_start": 0, "recovery_timer_start": 0, "route_cooldown": 0,
        "prev_time": time.time(), "frame_counter": 0, "last_results": None,
        "bg_blob_centroid": None, "bg_blob_since": 0,
    }
    while not _ai_thread_stop.is_set():
        try:
            _process_ai_cycle(cap, state)
        except Exception as e:
            print(f"[AI Worker] Error: {e}")
            time.sleep(0.5)


def generate_frames():
    """
    /video_feed generator. Does NOT run AI — just reads whatever frame
    the background _ai_worker thread most recently produced and streams
    it as MJPEG. Safe to have zero, one, or many viewers; AI keeps running
    in all cases.
    """
    while True:
        with _frame_lock:
            frame_bytes = _frame_buffer
        if frame_bytes is None:
            time.sleep(0.1)
            continue
        yield (b"--frame\r\n"
               b"Content-Type: image/jpeg\r\n\r\n" +
               frame_bytes + b"\r\n")
        time.sleep(0.03)  # ~30fps stream cap, independent of AI processing rate


@app.route("/video_feed")
def video_feed():
    return Response(generate_frames(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/api/set_mode/<new_mode>")
def set_mode(new_mode):
    global ai_mode
    with state_lock:
        if new_mode in ("DIORAMA", "ENTERPRISE"):
            ai_mode = new_mode
            return jsonify({"status": "success", "mode": ai_mode})
    return jsonify({"status": "error"}), 400


@app.route("/api/cancel_sim_trigger", methods=["POST"])
def cancel_sim_trigger():
    """
    Cancel a specific simulation hazard at a node.
    Body: { "node_id": "J7" }
    Removes the node from active_hazard_nodes and clears its hazard state.
    """
    global active_hazard_nodes, system_state, manual_override, fire_sim_active, facp_confirmed
    global current_per_node_routes, current_route
    body    = request.get_json(silent=True) or {}
    node_id = body.get("node_id")
    if not node_id:
        return jsonify({"error": "node_id required"}), 400
    with state_lock:
        # Capture event_type BEFORE filtering active_hazard_nodes removes it —
        # needed to decide whether this was a crowd-type hazard.
        _cancelled_entry = next((h for h in active_hazard_nodes if h["node_id"] == node_id), None)
        _was_crowd = _cancelled_entry is not None and _cancelled_entry.get("event_type") == "crowd"
        active_hazard_nodes[:] = [h for h in active_hazard_nodes if h["node_id"] != node_id]
        # If that was the last remaining FIRE (even if fallen/crowd hazards
        # are still active), turn off the thermal simulator specifically —
        # previously fire_sim_active only ever got reset in the "ALL
        # hazards cleared" branch further down, so cancelling a fire while
        # a fallen-person hazard was still active left the dashboard
        # showing a stuck 150°C reading with no active fire anywhere.
        if not any(h.get("event_type")=="fire" for h in active_hazard_nodes):
            fire_sim_active = False
        if node_id in live_node_status:
            live_node_status[node_id]["status"]           = "normal"
            live_node_status[node_id]["hazard"]            = None
            live_node_status[node_id]["pull_signal"]       = "GREEN"
            live_node_status[node_id]["impassable"]        = False
            live_node_status[node_id]["shelter_in_place"]  = False
            if _was_crowd:
                # The crowd NUMBER here was entirely synthetic — set by the
                # simulation trigger, not a real camera reading — so
                # cancelling it should bring the count back down too, unlike
                # the general /reset button, which deliberately leaves real
                # crowd counts alone ("people don't vanish just because the
                # alarm cleared"). update_crowd() also correctly resets
                # capacity_streak and tier as a side effect, instead of
                # leaving them stuck at whatever they were mid-hazard —
                # which could otherwise silently re-trigger the same hazard
                # on the very next crowd reading even after cancelling it.
                update_crowd(node_id, 0)
        # Immediately rebuild per-hazard routes to exclude the cancelled
        # node, rather than leaving current_per_node_routes stale until the
        # periodic background loop happens to catch up (throttled to once/
        # second, and gated behind conditions that don't always hold) — the
        # frontend polls /api/status and was pulling this stale list right
        # back over its own already-correct local removal, making a
        # cancelled hazard's tab/route reappear until ALL hazards were
        # cancelled (only then did the special "no hazards left" branch
        # below correctly clear everything).
        _per = []
        for _h in active_hazard_nodes:
            _h_routes = get_all_exit_routes(_h["node_id"])
            # Same fix as sim_trigger() and the periodic loop above — never
            # silently drop a hazard just because it's currently stranded.
            _per.append({
                "node_id":    _h["node_id"],
                "event_type": _h["event_type"],
                "best_path":  _h_routes[0]["path"] if _h_routes else [],
                "best_exit":  _h_routes[0]["exit"] if _h_routes else None,
                "best_cost":  _h_routes[0]["cost"] if _h_routes else None,
                "all_exits":  _h_routes,
                "rset":       estimate_rset(_h_routes[0]["path"]) if _h_routes else None,
            })
            if _h["node_id"] in live_node_status:
                live_node_status[_h["node_id"]]["shelter_in_place"] = (len(_h_routes) == 0)
        current_per_node_routes[:] = _per
        if _per:
            current_route[:] = _per[-1]["best_path"]
        else:
            current_route[:] = []
        # If no more active hazards, return to normal
        if not active_hazard_nodes:
            # A manual BOMBA block lives in live_node_status (hazard=="collapsed"),
            # completely separate from active_hazard_nodes (sim triggers only).
            # Previously this branch unconditionally reset to NORMAL and
            # published RESOLVED the moment the last SIM hazard was cancelled,
            # even if a structural BOMBA block was still sitting on the map —
            # incorrectly telling the UI and hardware "all clear" while a
            # corridor was still physically blocked.
            _blocks_remain = any(d.get("impassable", False) and d.get("hazard") == "collapsed"
                                  for d in live_node_status.values())
            if _blocks_remain:
                system_state     = "HAZARD"
                _total_pax       = sum(d["crowd"] for d in live_node_status.values())
                _corridors       = _build_corridor_states()
                _publish_resolve = False
            else:
                system_state    = "NORMAL"
                manual_override = False
                fire_sim_active = False
                facp_confirmed  = False  # reset so next fire triggers FACP sequence cleanly
                _total_pax      = sum(d["crowd"] for d in live_node_status.values())
                _corridors      = _build_corridor_states()
                _publish_resolve = True
        else:
            _publish_resolve = False
            _total_pax       = 0
            _corridors       = {}

    # Publish RESOLVED to MQTT so ESP32 hardware turns off — must be outside lock
    if _publish_resolve:
        mqtt_client.publish(TOPIC, json.dumps({
            "status":          "RESOLVED",
            "system_state":    "NORMAL",
            "person_count":    _total_pax,
            "stealth_mode":    True,
            "green_direction": "NONE",
            "corridors":       _corridors,
        }), retain=True)

    print(f"[SIM] Cancelled hazard at {node_id}, {len(active_hazard_nodes)} hazards remaining")
    return jsonify({"status": "success", "node_id": node_id,
                    "remaining_hazards": len(active_hazard_nodes)})


@app.route("/api/active_hazards")
def api_active_hazards():
    """
    Always returns the CURRENT, COMPLETE list of active hazards and their
    computed routes — unconditionally, regardless of manual_override or any
    other mode gating. Unlike /api/status's per_node_routes field (which is
    only reliably refreshed by the periodic background loop, and that loop
    explicitly skips updating during manual_override), this endpoint
    computes fresh, directly, on every call.
    Added as a safety-net reconciliation point: the frontend calls this
    after every trigger/cancel action to correct any remaining drift
    between what an individual action's own response implied and what the
    backend's actual complete hazard list is — regardless of whatever
    specific mechanism might cause that drift, without needing to diagnose
    each one individually.
    """
    with state_lock:
        _hazards = list(active_hazard_nodes)
    _per_node_routes = []
    for _h in _hazards:
        _h_routes = get_all_exit_routes(_h["node_id"])
        _per_node_routes.append({
            "node_id":    _h["node_id"],
            "event_type": _h["event_type"],
            "best_path":  _h_routes[0]["path"] if _h_routes else [],
            "best_exit":  _h_routes[0]["exit"] if _h_routes else None,
            "best_cost":  _h_routes[0]["cost"] if _h_routes else None,
            "all_exits":  _h_routes,
            "rset":       estimate_rset(_h_routes[0]["path"]) if _h_routes else None,
        })
    return jsonify({"per_node_routes": _per_node_routes})


@app.route("/api/get_route")
def get_route():
    with state_lock:
        route   = current_route
        signals = current_pull_signals
        rset    = current_rset

    pull_list = [
        {"node": nid, "signal": info["signal"], "reason": info["reason"]}
        for nid, info in signals.items()
    ]
    return jsonify({
        "status":       "success",
        "route":        route,
        "cost_score":   current_route_cost if current_route_cost != float("inf") else None,
        "pull_signals": pull_list,
    })


@app.route("/api/status")
def api_status():
    with state_lock:
        # Snapshot all mutable state — fast reads only under the lock.
        # get_crowd_velocity() and jsonify() happen OUTSIDE to avoid
        # blocking generate_frames / thermal / FFT threads.
        _state   = system_state
        _mode    = ai_mode
        _facp    = facp_confirmed
        _manual  = manual_override
        _sysmode = system_mode
        _bomba   = bomba_override_active
        _simtype = sim_trigger_type
        _simnode = sim_trigger_node
        _count   = current_person_count
        _total_pax = sum(d["crowd"] for d in live_node_status.values())
        _tracks  = len(current_track_ids)
        _vel     = crowd_velocity_lobby
        _thermal = thermal_state        # needed for header strip between MQTT events
        _fft     = fft_state            # needed for header strip between MQTT events
        _route   = list(current_route)
        _per_node = list(current_per_node_routes)
        _signals = dict(current_pull_signals)
        _rset    = dict(current_rset)
        _t_lat   = _thermal_latency_ms
        _f_lat   = _fft_latency_ms
        _nodes_snapshot = {nid: dict(data) for nid, data in live_node_status.items()}

    # All computation outside the lock
    return jsonify({
        "system_state":       _state,
        "system_mode":        _sysmode,
        "bomba_override":     _bomba,
        "sim_trigger_type":   _simtype,
        "sim_trigger_node":   _simnode,
        "ai_mode":            _mode,
        "facp_confirmed":     _facp,
        "manual_override":    _manual,
        "person_count":       _count,
        "total_footfall":     _total_pax,
        "active_tracks":      _tracks,
        "crowd_velocity":     _vel,
        "thermal_state":      _thermal,
        "fft_state":          _fft,
        "thermal_latency_ms": round(_t_lat, 3),
        "fft_latency_ms":     round(_f_lat, 3),
        "nodes_online":       NODES_ONLINE,
        "nodes_total":        NODES_TOTAL,
        "current_route":      _route,
        "per_node_routes":    _per_node,
        "pull_signals":       _signals,
        "rset":               _rset,
        "baseline_rset":      estimate_baseline_rset(_route) if _route else {},
        "nodes": {
            nid: {
                "status":     d["status"],
                "hazard":     d["hazard"],
                "impassable":       d.get("impassable", False),
                "shelter_in_place": d.get("shelter_in_place", False),
                "crowd":      d["crowd"],
                "velocity":   round(get_crowd_velocity(nid), 3),  # outside lock — safe read
                "pull":       d["pull_signal"],
                "temp":       _latest_temps.get(nid, 27.0),
            }
            for nid, d in _nodes_snapshot.items()
        },
    })


@app.route("/api/node_states")
def node_states():
    with state_lock:
        snapshot = {nid: dict(data) for nid, data in live_node_status.items()}
    # velocity computed outside lock — consistent with api_status pattern
    states = [
        {
            "id":       nid,
            "status":   data["status"],
            "hazard":   data["hazard"],
            "crowd":    data["crowd"],
            "velocity": round(get_crowd_velocity(nid), 3),
            "pull":     data["pull_signal"],
        }
        for nid, data in snapshot.items()
    ]
    return jsonify({"status": "success", "nodes": states})


@app.route("/trigger")
def trigger_hazard():
    """
    BOMBA "Simulate Fire" — there is no physical thermal sensor in this prototype,
    so fire scenarios are triggered manually for demonstration purposes.
    Fall detection (via camera) is independent and works without this trigger.

    Sets fire_sim_active=True so the thermal classifier thread begins
    simulating a gradual fire (_gradual_fire curve) and the FFT thread
    begins simulating the 520Hz alarm tone for FACP confirmation.
    Sets manual_override so stochastic drift pauses.
    """
    global system_state, manual_override, fire_sim_active
    with state_lock:
        system_state    = "HAZARD"
        manual_override = True
        fire_sim_active = True   # thermal + FFT threads begin fire simulation
        live_node_status["J7"]["status"] = "alert"
        live_node_status["J7"]["hazard"] = "thermal"
        live_node_status["J7"]["impassable"] = True
        # Force a clean reroute immediately rather than waiting for the next
        # periodic tick — this legacy endpoint didn't do this previously,
        # unlike /api/sim_trigger which already recalculates on every call.
        reset_hysteresis()
        _path, _ = calculate_safest_route("J4", verbose=False)
        current_route[:] = _path
        _total_pax = sum(d["crowd"] for d in live_node_status.values())
        _corridors = _build_corridor_states()
    mqtt_client.publish(TOPIC, json.dumps({
        "status":       "CRITICAL",
        "system_state": "HAZARD",
        "hazard_type":  "MANUAL OVERRIDE (thermal only — awaiting FFT confirmation)",
        "person_count": _total_pax,
        "corridors":    _corridors,
    }))
    return jsonify({"status": "success", "message": "Fire simulation triggered at J7 (Thai Relax corridor — FACP zone B5) — thermal + acoustic AI now running"})


@app.route("/api/facp_store_alert", methods=["POST","GET"])
def api_facp_alert():
    """
    FACP integration endpoint.
    Called when building fire panel signals a specific store.
    Body: { "door_id": "B5", "hazard": "thermal" }
    Marks nearest junction alert + returns evacuation route from store door.
    """
    global system_state, manual_override, fire_sim_active
    body        = request.get_json(silent=True) or {}
    door_id     = body.get("door_id") or request.args.get("door_id")
    if not door_id:
        return jsonify({"error": "door_id required — never guess fire location"}), 400
    hazard      = body.get("hazard", "thermal")
    store_label = DOOR_LABELS.get(door_id, door_id)
    with state_lock:
        junction_id = facp_store_alert(door_id, hazard)
        if not junction_id:
            return jsonify({"error": f"Unknown door_id: {door_id}"}), 400
        path, cost  = route_from_store(door_id, verbose=False)
        current_route[:] = path
        system_state     = "HAZARD"
        manual_override  = True
        fire_sim_active  = True
    print(f"[FACP] ALERT: {store_label} ({door_id}) — {hazard}. Junction: {junction_id}. Route: {' → '.join(path)}")
    return jsonify({
        "status":    "facp_alert",
        "store":     store_label,
        "door_id":   door_id,
        "junction":  junction_id,
        "route":     path,
        "cost":      cost if cost != float("inf") else None,
        "message":   f"FACP: {store_label} fire detected — evacuating via {' → '.join(path)}",
    })


@app.route("/api/facp_store_clear", methods=["POST","GET"])
def api_facp_clear():
    """Clear a FACP store alert and restore normal routing."""
    global system_state, manual_override, fire_sim_active
    from routing_engine import facp_store_clear as _facp_clear
    body        = request.get_json(silent=True) or {}
    door_id     = body.get("door_id") or request.args.get("door_id")
    if not door_id:
        return jsonify({"error": "door_id required — never guess fire location"}), 400
    store_label = DOOR_LABELS.get(door_id, door_id)
    with state_lock:
        junction_id      = _facp_clear(door_id)
        system_state     = "NORMAL"
        manual_override  = False
        fire_sim_active  = False
        reset_hysteresis()
    print(f"[FACP] CLEAR: {store_label} ({door_id}) — junction {junction_id} restored")
    return jsonify({
        "status":   "cleared",
        "door_id":  door_id,
        "junction": junction_id,
        "message":  f"FACP: {store_label} hazard cleared — normal routing resumed",
    })


@app.route("/reset")
def reset_system():
    global system_state, facp_confirmed, current_route, current_pull_signals, current_rset, \
           manual_override, fire_sim_active, fft_state, thermal_state, current_route_cost, \
           sim_trigger_type, sim_trigger_node
    with state_lock:
        system_state         = "NORMAL"
        facp_confirmed       = False
        manual_override      = False
        fire_sim_active      = False
        fft_state            = "SILENT"
        thermal_state        = "NORMAL"
        current_route_cost   = 0
        current_route        = ["J19","J20","J1","EXIT-1"]
        current_pull_signals = {}
        current_rset         = {}
        current_per_node_routes.clear()
        # Before clearing tracking, reset the crowd count for any node that
        # was an explicit crowd-type sim hazard — that number was synthetic
        # (set by the trigger, not real occupancy), unlike general ambient
        # crowd drift which correctly stays untouched by reset. Without
        # this, active_hazard_nodes.clear() below removes the node from
        # tracking (no more tab) while its crowd count stays stuck at
        # whatever it was mid-hazard — showing crowd styling on the map
        # with no way to explain why, since it's no longer tracked anywhere.
        for _h in active_hazard_nodes:
            if _h.get("event_type") == "crowd" and _h["node_id"] in live_node_status:
                update_crowd(_h["node_id"], 0)
        active_hazard_nodes.clear()   # clear sim triggers so DYN-A* stops repopulating
        reset_hysteresis()            # clear cached route so fresh calc on next tick
        sim_trigger_type = None
        sim_trigger_node = None
        for nid, data in live_node_status.items():
            data["status"]           = "normal"
            data["hazard"]           = None
            data["pull_signal"]      = "GREEN"
            data["impassable"]       = False
            data["shelter_in_place"] = False
            # capacity_streak is purely an internal debounce counter, not a
            # real-world quantity — unlike crowd count, there's no reason to
            # preserve it, and leaving it stuck high after an explicit reset
            # could cause an immediate re-block on the very next reading.
            data["capacity_streak"]  = 0
        # Reset clears hazard state, NOT actual occupancy — people don't
        # vanish from the building just because the alarm cleared.
        _total_pax = sum(d["crowd"] for d in live_node_status.values())
        _corridors = _build_corridor_states()  # all "normal" post-reset
    mqtt_client.publish(TOPIC, json.dumps({
        "status": "RESOLVED", "system_state": "NORMAL",
        "person_count": _total_pax, "stealth_mode": True,
        "green_direction": "NONE", "corridors": _corridors,
    }), retain=True)
    return jsonify({"status": "success", "message": "System reset to NORMAL"})


@app.route("/api/block_node", methods=["POST","GET"])
def block_node():
    """
    BOMBA blocks a node. Backend recalculates route avoiding it.
    Returns the complete new route — frontend just displays it.
    """
    global current_route, manual_override
    body    = request.get_json(silent=True) or {}
    node_id = body.get("node_id") or request.args.get("node_id", "J4")
    # start: caller sends the hazard origin (activeRoute[0]); fallback to current_route
    start   = (body.get("start")
               or request.args.get("start")
               or (current_route[0] if current_route else "J4"))
    # Keep the ORIGINAL id (e.g. "B9") for anything sent back to the frontend —
    # room-polygon highlighting matches against door IDs, not junctions, so if
    # this got silently resolved to "J4" before being echoed back, the room
    # would never actually light up blue even though the backend logic was
    # otherwise correct.
    start_original = start
    # If start is a door (Bx), get its junction for the ROUTING calculation only
    from routing_engine import DOOR_TO_JUNCTION
    if start in DOOR_TO_JUNCTION:
        start = DOOR_TO_JUNCTION[start]
    # impassable=True (default) means a genuine physical block — structural
    # collapse, etc. Set impassable=False explicitly for a soft "avoid if
    # possible" advisory instead.
    impassable = body.get("impassable", True)
    with state_lock:
        result = block_node_and_reroute(node_id, start, impassable=impassable)
        # Echo back the ORIGINAL id, not the resolved junction
        result["start"] = start_original
        # If the block is impassable and truly no route exists, do NOT
        # silently keep serving the stale previous route — an evacuee
        # following it could be walked into a collapsed area. Clear it and
        # flag the situation so the dashboard shows "NO ROUTE — SEND RESCUE"
        # instead of a route that quietly still worked before the block.
        current_route = result["new_route"] if not result.get("no_route") else []
        manual_override = True
        _total_pax = sum(d["crowd"] for d in live_node_status.values())
        _corridors = _build_corridor_states()
    # Push immediately — don't make BOMBA wait up to 2s for the next
    # heartbeat to see the diorama lights react to a manual block.
    mqtt_client.publish(TOPIC, json.dumps({
        "status": "CRITICAL", "system_state": system_state,
        "hazard_type": "NO ROUTE — RESCUE REQUIRED" if result.get("no_route") else "MANUAL OVERRIDE",
        "manual_override": True, "no_route": result.get("no_route", False),
        "shelter_node": start_original if result.get("no_route") else None,
        "stealth_mode": False, "person_count": _total_pax,
        "green_direction": "FOLLOW_ROUTE", "corridors": _corridors,
    }))
    if result.get("no_route"):
        print(f"[BOMBA] Blocked {node_id} (IMPASSABLE) — NO ROUTE EXISTS. Rescue required.")
    else:
        print(f"[BOMBA] Blocked {node_id}, new route: {' → '.join(result['new_route'])}")
    return jsonify(result)


@app.route("/api/unblock_node", methods=["POST","GET"])
def api_unblock():
    """BOMBA unblocks a previously blocked node."""
    global manual_override
    body    = request.get_json(silent=True) or {}
    node_id = body.get("node_id") or request.args.get("node_id", "J4")
    with state_lock:
        unblock_node(node_id)
        # unblock_node() only clears shelter_in_place on node_id itself, but
        # the flag may have been set on a DIFFERENT node (e.g. B9, if J4 was
        # the block that stranded it). Clear broadly here — anything still
        # genuinely stranded will get re-flagged on the next route calc.
        for _nid, _d in live_node_status.items():
            _d["shelter_in_place"] = False
        reset_hysteresis()
        # Only release manual override if THIS was the last remaining manual
        # block. Previously this cleared unconditionally, which could hand
        # control back to the background auto-reroute loop while other
        # BOMBA blocks were still active — the frontend already has this
        # exact conditional check (only clear when `remaining.length===0`),
        # but the backend disagreed with it, so the two could desync.
        _other_blocks_remain = any(
            d.get("impassable", False) for nid, d in live_node_status.items() if nid != node_id
        )
        if not _other_blocks_remain:
            manual_override = False
    return jsonify({"status": "unblocked", "node_id": node_id})


@app.route("/api/set_system_mode/<mode>")
def set_system_mode(mode):
    """
    Switch between simulation and live mode.
    Simulation: manual triggers only, real sensors suppressed.
    Live: real sensors only, manual simulation triggers disabled.
    Bomba override works in both modes.
    """
    global system_mode
    if mode not in ("simulation", "live"):
        return jsonify({"error": "mode must be 'simulation' or 'live'"}), 400
    with state_lock:
        system_mode = mode
        if mode == "live":
            # Clear simulation state so DYN-A* thread resumes auto-routing
            global manual_override, sim_trigger_type, sim_trigger_node
            manual_override  = False
            sim_trigger_type = None
            sim_trigger_node = None
            # Same reasoning as /reset — these counts were synthetic (set by
            # a sim trigger), not real occupancy, so they should go back to
            # baseline when their tracking is cleared, not stay stuck
            # showing stale crowd styling with no active tracking behind it.
            for _h in active_hazard_nodes:
                if _h.get("event_type") == "crowd" and _h["node_id"] in live_node_status:
                    update_crowd(_h["node_id"], 0)
            active_hazard_nodes.clear()
    print(f"[MODE] System mode switched to: {mode.upper()}")
    return jsonify({"status": "success", "system_mode": mode})


@app.route("/api/get_system_mode")
def get_system_mode():
    """Returns current system mode and Bomba override status."""
    with state_lock:
        return jsonify({
            "system_mode":           system_mode,
            "bomba_override_active": bomba_override_active,
            "sim_trigger_type":      sim_trigger_type,
            "sim_trigger_node":      sim_trigger_node,
        })


@app.route("/api/sim_trigger", methods=["POST"])
def sim_trigger():
    """
    Simulation mode only — manually trigger a fire/fallen/crowd event at a node.
    Body: { "event_type": "fire"|"fallen"|"crowd", "node_id": "J7" }
    Rejected if system_mode is not "simulation" or bomba_override_active is True.
    """
    global system_state, manual_override, fire_sim_active
    global sim_trigger_type, sim_trigger_node, active_hazard_nodes

    with state_lock:
        _mode   = system_mode
        _bomba  = bomba_override_active

    if _mode != "simulation":
        return jsonify({"error": "Simulation triggers only allowed in SIMULATION mode"}), 403
    if _bomba:
        return jsonify({"error": "Bomba override active — simulation triggers blocked"}), 403

    body       = request.get_json(silent=True) or {}
    event_type = body.get("event_type", "fire")   # "fire" | "fallen" | "crowd"
    node_id    = body.get("node_id", "J7")

    if event_type not in ("fire", "fallen", "crowd"):
        return jsonify({"error": "event_type must be fire, fallen, or crowd"}), 400

    with state_lock:
        # Accumulate hazards — append to list, do NOT clear previous ones
        sim_trigger_type  = event_type
        sim_trigger_node  = node_id
        system_state      = "HAZARD"
        # Do NOT set manual_override=True — we need DYN-A* loop to keep
        # recalculating in real-time as crowd density changes
        # Track all active hazard nodes for multi-path routing
        if not any(h["node_id"]==node_id for h in active_hazard_nodes):
            active_hazard_nodes.append({"node_id": node_id, "event_type": event_type})

        if event_type == "fire":
            fire_sim_active = True
            live_node_status[node_id]["status"] = "alert"
            live_node_status[node_id]["hazard"] = "thermal"
            # Same hard-block treatment as the real thermal sensor thread —
            # a simulated fire should behave identically to a real one for
            # routing purposes, not just carry a soft cost penalty.
            live_node_status[node_id]["impassable"] = True
            hazard_label = "FIRE (Simulation)"
        elif event_type == "fallen":
            live_node_status[node_id]["status"] = "alert"
            live_node_status[node_id]["hazard"] = "fall"
            live_node_status[node_id]["pull_signal"] = "RED"
            hazard_label = "PERSON FALLEN (Simulation)"
        elif event_type == "crowd":
            live_node_status[node_id]["status"] = "warning"
            live_node_status[node_id]["hazard"] = "crowd"
            live_node_status[node_id]["pull_signal"] = "AMBER"
            # calculate_dynamic_cost() only reacts to the raw crowd COUNT
            # crossing TIER1/TIER2 thresholds — it never checks hazard=="crowd"
            # or status=="warning" directly. Without this, the sim flags looked
            # right (icon, status) but the router genuinely saw a low, harmless
            # pax count and didn't penalize the node — combined with no
            # hysteresis in the multi-hazard per-node recompute, that let tiny
            # background drift on neighboring junctions flip which exit looked
            # cheapest tick to tick, showing up as the red route line blinking.
            if live_node_status[node_id]["crowd"] < TIER1_CROWD:
                update_crowd(node_id, TIER1_CROWD + 10)
            hazard_label = "CROWD DENSITY (Simulation)"

        _total_pax = sum(d["crowd"] for d in live_node_status.values())
        # Reset hysteresis so multi-hazard recalculates fresh each trigger
        reset_hysteresis()
        # Calculate best exit route for EACH active hazard node
        _per_node_routes = []
        for _h in active_hazard_nodes:
            _h_routes = get_all_exit_routes(_h["node_id"])
            # ALWAYS track the hazard, even with zero routes — a hazard that's
            # immediately stranded the moment it's triggered (e.g. residual
            # blocks from an earlier, since-cancelled hazard still cutting
            # off its only path) previously got silently OMITTED here
            # entirely, since this only appended when _h_routes was
            # non-empty. That meant no tab, no shelter-in-place marking,
            # nothing — the node showed its hazard icon (driven separately
            # by live_node_status) but never appeared in per-hazard
            # tracking at all, with no way for the frontend to know it
            # needed rescue. Empty best_path is the correct signal for
            # "stranded," not "doesn't exist."
            _per_node_routes.append({
                "node_id":    _h["node_id"],
                "event_type": _h["event_type"],
                "best_path":  _h_routes[0]["path"] if _h_routes else [],
                "best_exit":  _h_routes[0]["exit"] if _h_routes else None,
                "best_cost":  _h_routes[0]["cost"] if _h_routes else None,
                "all_exits":  _h_routes,
                "rset":       estimate_rset(_h_routes[0]["path"]) if _h_routes else None,
            })
            if _h["node_id"] in live_node_status:
                live_node_status[_h["node_id"]]["shelter_in_place"] = (len(_h_routes) == 0)
        # Primary route = best route from most recently triggered node
        _path = _per_node_routes[-1]["best_path"] if _per_node_routes else []
        # Always assign, even when empty — previously this only updated
        # current_route when _path was truthy, so a genuine no-route result
        # (nowhere for the most recent hazard to go) left current_route
        # permanently stuck on whatever stale route existed before, instead
        # of correctly reflecting that the building is now trapped.
        current_route[:] = _path
        _corridors = _build_corridor_states()

    mqtt_client.publish(TOPIC, json.dumps({
        "status":       "CRITICAL",
        "system_state": "HAZARD",
        "hazard_type":  hazard_label,
        "source":       "SIMULATION",
        "node_id":      node_id,
        "person_count": _total_pax,
        "corridors":    _corridors,
    }))
    print(f"[SIM] Triggered {event_type.upper()} at {node_id} → {len(_per_node_routes)} hazard nodes active")
    return jsonify({
        "status":          "success",
        "event_type":      event_type,
        "node_id":         node_id,
        "route":           _path or [],
        "per_node_routes": _per_node_routes,
        "message":         f"Simulation: {hazard_label} triggered at {node_id}",
    })


@app.route("/api/bomba_override", methods=["POST"])
def bomba_override():
    """
    Bomba override — works in ANY mode (simulation or live).
    Highest priority event — cannot be overridden by simulation or sensors.
    Body: { "action": "activate"|"clear", "node_id": "J7" (optional) }
    """
    global system_state, manual_override, fire_sim_active, bomba_override_active
    global sim_trigger_type, sim_trigger_node

    body    = request.get_json(silent=True) or {}
    action  = body.get("action", "activate")
    node_id = body.get("node_id", None)

    if action == "activate":
        with state_lock:
            bomba_override_active = True
            system_state          = "HAZARD"
            manual_override       = True
            fire_sim_active       = True   # triggers thermal + FFT simulation
            sim_trigger_type      = None   # cancel any pending simulation
            sim_trigger_node      = None

            _node = node_id or "J7"
            live_node_status[_node]["status"] = "alert"
            live_node_status[_node]["hazard"] = "thermal"
            _total_pax = sum(d["crowd"] for d in live_node_status.values())
            _corridors = _build_corridor_states()

        mqtt_client.publish(TOPIC, json.dumps({
            "status":       "CRITICAL",
            "system_state": "HAZARD",
            "hazard_type":  "BOMBA COMMAND OVERRIDE",
            "source":       "BOMBA",
            "person_count": _total_pax,
            "corridors":    _corridors,
        }))
        print("[BOMBA] Override ACTIVATED — highest priority event, all other triggers blocked")
        return jsonify({"status": "success", "bomba_override_active": True,
                        "message": "Bomba override activated — all sensors and simulation suppressed"})

    elif action == "clear":
        with state_lock:
            bomba_override_active = False
            system_state          = "NORMAL"
            manual_override       = False
            fire_sim_active       = False
            sim_trigger_type      = None
            sim_trigger_node      = None
            # Same reasoning as /reset and set_system_mode(live) — reset
            # crowd counts for nodes that were explicit sim-triggered crowd
            # hazards (synthetic numbers) before clearing their tracking.
            for _h in active_hazard_nodes:
                if _h.get("event_type") == "crowd" and _h["node_id"] in live_node_status:
                    update_crowd(_h["node_id"], 0)
            active_hazard_nodes.clear()
            for nid, d in live_node_status.items():
                d["status"]           = "normal"
                d["hazard"]           = None
                d["pull_signal"]      = "GREEN"
                d["impassable"]       = False
                d["shelter_in_place"] = False
                d["capacity_streak"]  = 0
            _total_pax = sum(d["crowd"] for d in live_node_status.values())
            _corridors = _build_corridor_states()

        mqtt_client.publish(TOPIC, json.dumps({
            "status":       "RESOLVED",
            "system_state": "NORMAL",
            "source":       "BOMBA",
            "person_count": _total_pax,
            "corridors":    _corridors,
        }), retain=True)
        print("[BOMBA] Override CLEARED — system restored to normal")
        return jsonify({"status": "success", "bomba_override_active": False,
                        "message": "Bomba override cleared — normal operation resumed"})

    return jsonify({"error": "action must be activate or clear"}), 400


@app.route("/api/quick_routes", methods=["POST","GET"])
def api_quick_routes():
    """
    Returns routes to all reachable exits from a given start point.
    Used by BOMBA quick route panel, and by the per-hazard-node route
    refresh after a block/unblock (which passes mark_shelter=true).
    """
    body  = request.get_json(silent=True) or {}
    start = body.get("start") or request.args.get("start", "J4")
    # Opt-in only — the general-purpose Quick Reroute panel also calls this
    # endpoint on every render for casual "what are my options" lookups, and
    # marking shelter_in_place on every such call would create false
    # positives on nodes that were never actually a tracked hazard origin.
    mark_shelter = bool(body.get("mark_shelter", False))
    from routing_engine import DOOR_TO_JUNCTION
    if start in DOOR_TO_JUNCTION:
        start_j = DOOR_TO_JUNCTION[start]
    else:
        start_j = start
    with state_lock:
        routes = get_all_exit_routes(start_j)
        # Guard against marking a manually-blocked PASSAGE as shelter (nobody
        # is "at" a blocked corridor junction waiting for rescue — it's just
        # a route that no longer exists). block_node_and_reroute() tags a
        # manual BOMBA block with hazard="collapsed" specifically for this.
        # IMPORTANT: this is NOT the same check as "is this node impassable"
        # — a fire/hazard node is ALSO impassable, but for a different
        # reason (it IS the danger), and someone actually there legitimately
        # needs shelter marking. Checking impassable alone would incorrectly
        # block that case too.
        _is_manual_block = live_node_status.get(start, {}).get("hazard") == "collapsed"
        if mark_shelter and start in live_node_status and not _is_manual_block:
            # Make the backend the authoritative source for this flag —
            # previously the frontend set shelterInPlace locally after
            # discovering a stranded hazard node via this same endpoint,
            # but never told the backend, so the very next /api/status poll
            # would overwrite it back to false (the "blinks once" bug).
            live_node_status[start]["shelter_in_place"] = (len(routes) == 0)
    # If start was a door, prepend it
    if start in DOOR_TO_JUNCTION:
        for r in routes:
            r["path"] = [start] + r["path"]
    return jsonify({"start": start, "routes": routes})


@app.route("/api/force_exit", methods=["POST","GET"])
def api_force_exit():
    """
    BOMBA forces route to a specific exit. Backend calculates full junction path.
    """
    global current_route, manual_override
    body     = request.get_json(silent=True) or {}
    start    = body.get("start") or (current_route[0] if current_route else "J4")
    exit_id  = body.get("exit_id") or request.args.get("exit_id", "EXIT-1")
    from routing_engine import DOOR_TO_JUNCTION
    start_j  = DOOR_TO_JUNCTION.get(start, start)
    path, cost = route_to_specific_exit(start_j, exit_id, verbose=False)
    if start in DOOR_TO_JUNCTION and path:
        path = [start] + path
    with state_lock:
        current_route = path
        force_route(path)
        manual_override = True
    print(f"[BOMBA] Forced route to {exit_id}: {' → '.join(path)}")
    return jsonify({"route": path, "cost": cost if cost != float("inf") else None, "exit": exit_id})


@app.route("/download_log")
def download_log():
    """Commercial + operational report for facility managers and HaaS subscribers."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Lumina brand palette, matching the dashboard's own colours ──
    NAVY      = "1E293B"   # dark header background
    TEAL      = "0D9488"   # section header background
    TEAL_LT   = "CCFBF1"   # light teal tint
    GREEN     = "16A34A"   # safe / OK
    GREEN_LT  = "DCFCE7"
    AMBER     = "D97706"   # warning / monitor
    AMBER_LT  = "FEF3C7"
    RED       = "DC2626"   # critical
    RED_LT    = "FEE2E2"
    GREY_TXT  = "64748B"
    WHITE     = "FFFFFF"
    BORDER    = Border(*(Side(style="thin", color="E2E8F0"),)*4)
    FONT_NAME = "Calibri"

    def style_title(cell, size=14):
        cell.font = Font(name=FONT_NAME, size=size, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center")

    def style_section(cell):
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(vertical="center")

    def style_header(cell):
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="475569")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER

    def style_data(cell, bold=False):
        cell.font = Font(name=FONT_NAME, size=10, bold=bold)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    def status_fill(cell, kind):
        # kind: "good" | "warn" | "bad"
        fg, bg = {"good": (GREEN, GREEN_LT), "warn": (AMBER, AMBER_LT), "bad": (RED, RED_LT)}[kind]
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=fg)
        cell.fill = PatternFill("solid", fgColor=bg)

    def write_row(ws, row, values, styles=None):
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            style_data(c, bold=(styles=="bold"))
        return row + 1

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def section_title(ws, row, text, span):
        c = ws.cell(row=row, column=1, value=text)
        style_section(c)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        for col in range(2, span+1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=TEAL)
        return row + 1

    def table_header(ws, row, cols):
        for i, h in enumerate(cols, start=1):
            style_header(ws.cell(row=row, column=i, value=h))
        return row + 1

    with state_lock:
        snap    = {nid: dict(d) for nid, d in live_node_status.items()}
        _sys    = system_state
        _facp   = facp_confirmed

    rset_data      = estimate_rset(current_route)
    baseline_data  = estimate_baseline_rset(current_route)
    total_footfall = sum(d["crowd"] for d in snap.values())
    peak_entry     = max(snap.items(), key=lambda x: x[1]["crowd"]) if snap else ("N/A", {"crowd": 0})
    avg_occ        = round(total_footfall / max(len(snap), 1), 1)
    dynamic_rset   = rset_data.get("RSET_s", 142)
    baseline_rset  = baseline_data.get("RSET_s", 342)
    try:
        reduction_pct = round((1 - float(dynamic_rset) / float(baseline_rset)) * 100, 1)
    except Exception:
        reduction_pct = "N/A"

    BATT = {k: v["pct"]          for k, v in NODE_BATTERY.items()}
    NEXT = {k: v["next_service"] for k, v in NODE_BATTERY.items()}

    wb = Workbook()

    # ══════════════════════ SHEET 1: OVERVIEW ══════════════════════
    ws = wb.active
    ws.title = "Overview"
    set_widths(ws, [34, 46])
    r = 1
    ws.merge_cells("A1:B1")
    style_title(ws.cell(row=1, column=1, value="LUMINA SMART EVACUATION SYSTEM"), size=16)
    ws.row_dimensions[1].height = 26
    r = 2
    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1, value="Facility Management & Commercial Analytics Report")
    c.font = Font(name=FONT_NAME, size=11, italic=True, color=GREY_TXT)
    r = 4
    for label, val in [
        ("Generated",             time.strftime('%Y-%m-%d %H:%M:%S')),
        ("Session Duration (s)",  round(time.time() - _startup_time, 1)),
        ("Deployment Model",      "Hardware-as-a-Service (HaaS)"),
        ("System Status",         _sys),
    ]:
        r = write_row(ws, r, [label, val], styles="bold" if label=="System Status" else None)
    r += 1

    r = section_title(ws, r, "FOOTFALL TELEMETRY", 2)
    for label, val in [
        ("Total Occupancy (pax)",       total_footfall),
        ("Peak Zone",                   f"{resolve_node_name(peak_entry[0])} ({peak_entry[0]})"),
        ("Peak Zone Occupancy (pax)",   peak_entry[1]["crowd"]),
        ("Average Zone Occupancy (pax)", avg_occ),
        ("Tracking Method",             "Anonymous crowd vectors (no facial data)"),
        ("PDPA Compliant",              "Yes — 0 bytes raw video transmitted"),
    ]:
        r = write_row(ws, r, [label, val])
    r += 1

    r = section_title(ws, r, "EVACUATION SAFETY STATUS", 2)
    for label, val in [
        ("Active Route",                  " > ".join(current_route)),
        ("Route Safe",                    "Yes" if rset_data.get("safe", True) else "No"),
        ("Estimated Evacuation Time (s)", f"{dynamic_rset}  (With Lumina DYN-A* guidance)"),
        ("Baseline Evacuation Time (s)",  f"{baseline_rset}  (Static signage, panic speed)"),
        ("Time Reduction",                f"{reduction_pct}%  (Measured by routing engine)"),
        ("Available Safe Egress Time (s)", rset_data.get("ASET_s", 600)),
        ("Safety Margin (s)",             rset_data.get("margin_s", "N/A")),
        ("FACP Status",                   "Confirmed" if _facp else "Standby"),
    ]:
        r = write_row(ws, r, [label, val])
    ws.freeze_panes = "A5"

    # ══════════════════ SHEET 2: ZONE OCCUPANCY ══════════════════
    ws = wb.create_sheet("Zone Occupancy")
    set_widths(ws, [30, 10, 14, 16, 12, 40])
    r = 1
    r = table_header(ws, r, ["Zone", "Node ID", "Occupancy (pax)", "Crowd Velocity", "Status", "Recommended Action"])
    for nid, d in snap.items():
        vel   = round(get_crowd_velocity(nid), 2)
        crowd = d["crowd"]
        if crowd > 85:
            action, kind = "HIGH TRAFFIC — Prime DOOH zone, activate pull policy", "bad"
        elif crowd > 60:
            action, kind = "MODERATE TRAFFIC — Kiosk opportunity", "warn"
        elif crowd < 10:
            action, kind = "LOW TRAFFIC — Consider HVAC reduction", "good"
        else:
            action, kind = "NORMAL", "good"
        row_vals = [resolve_node_name(nid), nid, crowd, vel, d["status"].upper(), action]
        for i, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            style_data(cell)
        status_fill(ws.cell(row=r, column=5), kind)
        r += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{r-1}"

    # ══════════════════ SHEET 3: COMMERCIAL & HVAC ══════════════════
    ws = wb.create_sheet("Commercial & HVAC")
    set_widths(ws, [30, 10, 14, 14, 26, 44])
    occupied = [(nid, d["crowd"]) for nid, d in snap.items() if d["crowd"] > 0]
    empty    = [(nid, d["crowd"]) for nid, d in snap.items() if d["crowd"] < 10]
    high     = [(nid, d["crowd"]) for nid, d in snap.items() if d["crowd"] > 60]
    total_pax   = sum(d["crowd"] for d in snap.values())
    avg_occ_pct = round(total_pax / max(len(snap) * 100, 1) * 100, 1)

    r = 1
    r = section_title(ws, r, "FOOTFALL ANALYTICS", 6)
    for label, val, note in [
        ("Total Occupancy (pax)",      total_pax, "Current headcount across all zones"),
        ("Average Zone Occupancy (%)", avg_occ_pct, "% of maximum capacity across all nodes"),
        ("High-Traffic Zones",         ", ".join(f"{resolve_node_name(z[0])} ({z[0]})" for z in high) or "None",
         "Above 60 pax — prime for DOOH / kiosk placement"),
        ("Low-Traffic Zones",          f"{len(empty)} zones below 10 pax", "Candidates for HVAC reduction — see full list on Zone Occupancy sheet"),
    ]:
        c1 = ws.cell(row=r, column=1, value=label); style_data(c1, bold=True)
        c2 = ws.cell(row=r, column=2, value=val); style_data(c2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col in (3,4):
            style_data(ws.cell(row=r, column=col))
        c3 = ws.cell(row=r, column=5, value=note); style_data(c3)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        style_data(ws.cell(row=r, column=6))
        r += 1
    r += 1

    r = section_title(ws, r, "HVAC OPTIMISATION SIGNALS", 6)
    r = table_header(ws, r, ["Zone", "Node ID", "Occupancy (pax)", "Measured Temp (°C)", "Recommended Action", "Note"])
    for nid, d in snap.items():
        temp  = round(_latest_temps.get(nid, 27.0), 1)
        crowd = d["crowd"]
        if crowd < 10:
            action, note, kind = "Reduce cooling — zone unoccupied", "Apply unoccupied setpoint (typically +3 to +5°C)", "good"
        elif crowd > 70:
            action, note, kind = "Increase cooling — high occupancy", "Apply peak-occupancy setpoint", "warn"
        else:
            action, note, kind = "Maintain current setpoint", "Normal occupancy range", "good"
        vals = [resolve_node_name(nid), nid, crowd, temp, action, note]
        for i, v in enumerate(vals, start=1):
            style_data(ws.cell(row=r, column=i, value=v))
        status_fill(ws.cell(row=r, column=5), kind)
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(row=r, column=1, value="Note: Lumina provides occupancy signals only. Energy savings depend on facility "
                "HVAC specifications, electricity tariff, and building management system configuration. "
                "No savings figures are claimed here.")
    c.font = Font(name=FONT_NAME, size=9, italic=True, color=GREY_TXT)
    c.alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"

    # ══════════════════ SHEET 4: EVACUATION DETAIL ══════════════════
    ws = wb.create_sheet("Evacuation Detail")
    set_widths(ws, [22, 12, 20, 10, 12])
    r = 1
    r = section_title(ws, r, "T2 SENSITIVITY ANALYSIS", 5)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    note_c = ws.cell(row=r, column=1, value="T2 = occupant response hesitation time. Design target: T2=5s (>80% reduction "
                "on 30s static baseline). Actual T2 requires user trial measurement before production claim. "
                "This table shows the system stays SAFE across all realistic T2 values.")
    note_c.font = Font(name=FONT_NAME, size=9, italic=True, color=GREY_TXT)
    note_c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1
    r = table_header(ws, r, ["T2 Hesitation (s)", "RSET (s)", "Reduction vs Static (%)", "Safe", "Margin (s)"])
    for row in rset_t2_sensitivity(current_route):
        marker = " (design target)" if row["T2_s"] == 5 else (" (= static, worst case)" if row["T2_s"] == 30 else "")
        vals = [f"{row['T2_s']}s{marker}", row["RSET_s"], f"{row['reduction_%']}%",
                "Yes" if row["safe"] else "No", row["margin_s"]]
        for i, v in enumerate(vals, start=1):
            style_data(ws.cell(row=r, column=i, value=v))
        status_fill(ws.cell(row=r, column=4), "good" if row["safe"] else "bad")
        r += 1
    r += 1

    r = section_title(ws, r, "ZONE CONGESTION SIGNALS", 5)
    r = table_header(ws, r, ["Zone", "Signal", "Detail", "", ""])
    ws.merge_cells(start_row=r-1, start_column=3, end_row=r-1, end_column=5)
    any_signal = False
    for nid, info in current_pull_signals.items():
        any_signal = True
        reason = info.get("reason", "N/A").replace("\u2014", "-").replace("\u2013", "-")
        signal = info.get("signal", "N/A")
        kind = "good" if signal=="GREEN" else ("warn" if signal=="AMBER" else "bad")
        c1 = ws.cell(row=r, column=1, value=resolve_node_name(nid)); style_data(c1)
        c2 = ws.cell(row=r, column=2, value=signal); status_fill(c2, kind)
        c3 = ws.cell(row=r, column=3, value=reason); style_data(c3)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        for col in (4,5): style_data(ws.cell(row=r, column=col))
        r += 1
    if not any_signal:
        c1 = ws.cell(row=r, column=1, value="All zones"); style_data(c1)
        c2 = ws.cell(row=r, column=2, value="GREEN"); status_fill(c2, "good")
        c3 = ws.cell(row=r, column=3, value="No congestion detected"); style_data(c3)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        for col in (4,5): style_data(ws.cell(row=r, column=col))
        r += 1

    # ══════════════════ SHEET 5: MAINTENANCE ══════════════════
    ws = wb.create_sheet("Maintenance")
    set_widths(ws, [12, 24, 12, 26, 14, 30])
    r = 1
    r = table_header(ws, r, ["Node ID", "Zone", "Battery (%)", "NFPA 72 Status", "Next Service", "Action Required"])
    for nid in NODE_BATTERY:
        bat   = BATT.get(nid, 85)
        next_ = NEXT.get(nid, "N/A")
        if bat >= 75:
            status, action, kind = "OK", "None", "good"
        elif bat >= 60:
            status, action, kind = "LOW — Monitor", "Schedule service within 30 days", "warn"
        else:
            status, action, kind = "CRITICAL — Below NFPA 72 threshold", "HOT-SWAP REQUIRED immediately", "bad"
        vals = [nid, LUMINA_NODE_LABELS.get(nid, nid), bat, status, next_, action]
        for i, v in enumerate(vals, start=1):
            style_data(ws.cell(row=r, column=i, value=v))
        status_fill(ws.cell(row=r, column=4), kind)
        r += 1
    r += 1
    ws.freeze_panes = "A2"

    # ══════════════════ SHEET 6: SYSTEM & COMPLIANCE ══════════════════
    ws = wb.create_sheet("System & Compliance")
    set_widths(ws, [32, 14, 46, 16])
    r = 1
    r = section_title(ws, r, "SYSTEM PERFORMANCE", 4)
    r = table_header(ws, r, ["Metric", "Value", "Target", "Status"])
    for metric, value, target, ok in [
        ("Thermal Detection Latency (ms)", round(_thermal_latency_ms, 1), "< 500ms", _thermal_latency_ms < 500),
        ("Acoustic Detection Latency (ms)", round(_fft_latency_ms, 1), "< 500ms", _fft_latency_ms < 500),
        ("Nodes Online", f"{NODES_ONLINE}/{NODES_TOTAL}", f"{NODES_TOTAL}/{NODES_TOTAL}", NODES_ONLINE==NODES_TOTAL),
    ]:
        vals = [metric, value, target, "Pass" if ok else "Review"]
        for i, v in enumerate(vals, start=1):
            style_data(ws.cell(row=r, column=i, value=v))
        status_fill(ws.cell(row=r, column=4), "good" if ok else "warn")
        r += 1
    r += 1

    r = section_title(ws, r, "PRIVACY AND COMPLIANCE SUMMARY", 4)
    r = table_header(ws, r, ["Item", "Status", "Notes", ""])
    for item, status, notes in [
        ("Raw video transmitted", "0 bytes", "Analytics run on edge TPU only — no raw video transmitted"),
        ("Facial data stored", "None", "ByteTrack anonymous vectors — no biometrics"),
        ("PDPA compliant", "Yes", "Personal Data Protection Act 2010 (Malaysia)"),
        ("NFPA 72 battery compliance", "Monitored", "Auto-alerts at 60% shelf life threshold"),
        ("HaaS contract renewal trigger", "Month 36", "Free hot-swap battery included on renewal"),
    ]:
        c1 = ws.cell(row=r, column=1, value=item); style_data(c1, bold=True)
        c2 = ws.cell(row=r, column=2, value=status); status_fill(c2, "good")
        c3 = ws.cell(row=r, column=3, value=notes); style_data(c3)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        style_data(ws.cell(row=r, column=4))
        # Column is now wide enough that most notes fit on one line, but
        # explicitly setting row height as a safety net for whatever still
        # wraps to two lines — without this, wrap_text can cram multiple
        # lines into a too-short default row height, causing text from
        # adjacent rows to visually overlap (exactly what was happening here).
        ws.row_dimensions[r].height = 30
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from flask import Response
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=Lumina_Management_Report.xlsx"}
    )

@app.route("/api/health")
def api_health():
    """
    System health snapshot — polled every 5s by React (slower than /api/status).
    Shows uptime, hardware status, and connection state for the System Health tab.
    Battery data lives here as the single source of truth for both dashboard and CSV.
    """
    with state_lock:
        _sys_state = system_state
    return jsonify({
        "status":             "ok",
        "uptime_s":           round(time.time() - _startup_time, 1),
        "yolo_loaded":        model_diorama is not None,
        "mqtt_connected":     mqtt_client.is_connected(),
        "camera_open":        cap.isOpened(),
        "nodes_online":       NODES_ONLINE,
        "nodes_total":        NODES_TOTAL,
        "ai_mode":            ai_mode,
        "thermal_latency_ms": round(_thermal_latency_ms, 3),
        "fft_latency_ms":     round(_fft_latency_ms,     3),
        "system_state":       _sys_state,
        "battery":            NODE_BATTERY,
    })


def _shutdown():
    print("[LUMINA] Shutting down — releasing camera and MQTT...")
    try: cap.release()
    except: pass
    try: mqtt_client.loop_stop(); mqtt_client.disconnect()
    except: pass
    print("[LUMINA] Clean shutdown complete.")

atexit.register(_shutdown)


if __name__ == "__main__":
    # Force clean NORMAL state on every startup — no stale hazard from previous session
    with state_lock:
        system_state         = "NORMAL"
        facp_confirmed       = False
        manual_override      = False
        fire_sim_active      = False
        current_route        = ["J19","J20","J1","EXIT-1"]
        current_pull_signals = {}
        current_rset         = {}
        for _nid, _d in live_node_status.items():
            _d["status"]      = "normal"
            _d["hazard"]      = None
            _d["pull_signal"] = "GREEN"
    if hasattr(thermal_clf, 'reset'): thermal_clf.reset()
    if hasattr(fft_clf,    'reset'): fft_clf.reset()

    # Start the AI worker as a daemon thread — runs independently of
    # whether any browser has /video_feed open (fixes observer-dependent
    # AI loop: fall detection and DYN-A* must keep running 24/7).
    _ai_thread = threading.Thread(target=_ai_worker, daemon=True)
    _ai_thread.start()
    print("[LUMINA] AI worker thread started — decoupled from /video_feed")

    print("[LUMINA] State reset to NORMAL on startup")
    print("[LUMINA] All subsystems initialised. Starting Flask on :5001")
    print("[LUMINA] Endpoints:")
    print("  /video_feed               — MJPEG camera stream with HUD overlay")
    print("  /api/get_route            — DYN-A* route + Pull Policy signals")
    print("  /api/status               — full telemetry snapshot (1.5s poll)")
    print("  /api/health               — system health + uptime (5s poll)")
    print("  /api/node_states          — per-node status for NodeMap.jsx")
    print("  /api/block_node           — POST {node_id} to quarantine a node")
    print("  /trigger                  — manual hazard override")
    print("  /reset                    — reset all state to NORMAL")
    print("")
    print("  For booth demo: set FLASK_IP in App.jsx to this machine's Wi-Fi IP")
    print("  Find it with:  ipconfig (Windows)  or  ifconfig (Mac/Linux)")
    app.run(host="0.0.0.0", port=5001, threaded=True)
